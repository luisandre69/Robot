# Import PAckage
import openpyxl

# Load Workbook
vk = openpyxl.load_workbook("Section18\Testdata.xlsx")

print(vk.sheetnames)
print("Active Sheet is " + vk.active.title)

# Create object of any sheet on which oyu want to work

sh = vk['Sheet1']

print(sh['A3'].value)
print(sh['B3'].value)

c1 = sh.cell(3, 1)
print(c1.value)

print(c1.row)
print(c1.column)

# Find Row and columns Having Data

rows = sh.max_row
columns = sh.max_column

print("Total Rows are - " + str(rows))
print("Total Columns are - " + str(columns))

for i in range(1, rows+1):
    for j in range(1, columns+1):
        c = sh.cell(i, j)
        print (c.value)


for r in sh['A1':'C3']:
    for c in r:
        print(c.value)
